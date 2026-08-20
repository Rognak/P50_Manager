"""Парсер Excel-файла со списком сотрудников.

Минимальный набор полей: ФИО, Email, Должность, Стаж работы.
Роль/грейд и заметки руководитель проставит вручную в карточке сотрудника
после импорта — там и удобнее, и нет риска ошибочного маппинга.
"""

from __future__ import annotations

import re
from datetime import date, timedelta
from io import BytesIO
from typing import Literal

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.employee import Employee
from app.schemas.employee import EmployeeImportRow

# --------- Колонки Excel (нормализованные заголовки → ключ) ----------------

COLUMN_ALIASES: dict[str, str] = {
    "фио": "full_name",
    "email": "email",
    "e-mail": "email",
    "должность": "position",
    "стаж работы": "tenure",
}


def _normalize_header(h: str) -> str:
    return (h or "").strip().lower()


# --------- Стаж работы → дата найма ----------------------------------------

_TENURE_RE = re.compile(
    r"(?:(?P<years>\d+)\s*(?:год|года|лет))?\s*"
    r"(?:и\s*)?"
    r"(?:(?P<months>\d+)\s*месяц)?",
    re.IGNORECASE,
)


def parse_tenure_to_hired_at(text: str | None) -> date | None:
    """Распарсить '2 года и 11 месяцев' → date(today - 35 месяцев).
    Если строка пустая или не распозналась — None."""
    if not text or not text.strip():
        return None
    m = _TENURE_RE.search(text)
    if not m or (not m.group("years") and not m.group("months")):
        return None
    years = int(m.group("years") or 0)
    months = int(m.group("months") or 0)
    total_days = years * 365 + months * 30
    if total_days == 0:
        return None
    return date.today() - timedelta(days=total_days)


# --------- Парсинг файла ----------------------------------------------------


async def parse_xlsx(
    file_bytes: bytes,
    session: AsyncSession,
    owner_id: int,
    department_id: int | None,
) -> list[EmployeeImportRow]:
    """Прочитать XLSX и вернуть список preview-строк (без записи в БД).

    `department_id` — отдел текущего DH, к которому будут привязаны все
    импортированные сотрудники (передаётся из UI; в Excel не парсим).

    Дедупликация: если у текущего DH уже есть Employee с таким email
    (или с таким full_name при пустом email) — `action='skip'`.
    """
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Не удалось прочитать XLSX: {e}")
    ws = wb.active
    if ws is None or ws.max_row < 2:
        return []

    # Заголовки
    headers_raw = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_to_key: dict[int, str] = {}
    for idx, h in enumerate(headers_raw):
        key = COLUMN_ALIASES.get(_normalize_header(str(h or "")))
        if key:
            header_to_key[idx] = key
    if "full_name" not in header_to_key.values():
        raise ValueError("В файле нет колонки 'ФИО'")

    # Существующие сотрудники этого DH — для дедупа
    existing = (
        await session.execute(
            select(Employee.full_name, Employee.email).where(
                Employee.owner_id == owner_id, Employee.kind == "employee"
            )
        )
    ).all()
    existing_emails = {e.lower() for _, e in existing if e}
    existing_names = {n.strip().lower() for n, _ in existing}

    out: list[EmployeeImportRow] = []
    for row_idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Пустые строки пропускаем
        if not row_values or all(v in (None, "") for v in row_values):
            continue

        # Маппим по индексам колонок
        data: dict[str, str | None] = {}
        for idx, key in header_to_key.items():
            if idx < len(row_values):
                v = row_values[idx]
                data[key] = str(v).strip() if v not in (None, "") else None

        warnings: list[str] = []
        full_name = data.get("full_name")
        if not full_name:
            out.append(
                EmployeeImportRow(
                    row=row_idx,
                    action="error",
                    full_name=None,
                    email=None,
                    position=None,
                    department_id=None,
                    hired_at=None,
                    warnings=[],
                    error="Не задано ФИО",
                )
            )
            continue

        email = data.get("email")
        position = data.get("position")
        hired_at = parse_tenure_to_hired_at(data.get("tenure"))

        # Дедуп
        action: Literal["create", "skip", "error"] = "create"
        if email and email.lower() in existing_emails:
            action = "skip"
            warnings.append("сотрудник с таким email уже существует")
        elif not email and full_name.strip().lower() in existing_names:
            action = "skip"
            warnings.append("сотрудник с таким ФИО уже существует (email пуст)")

        out.append(
            EmployeeImportRow(
                row=row_idx,
                action=action,
                full_name=full_name,
                email=email,
                position=position,
                department_id=department_id,  # из параметра, не из Excel
                hired_at=hired_at,
                warnings=warnings,
                error=None,
            )
        )

    return out
