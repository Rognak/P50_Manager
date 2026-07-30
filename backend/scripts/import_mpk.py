"""Импорт методики МПК из xlsx (листы: Модель ПК, Шкала оценки,
Новое! Профилирование ПК, Карта обучения).

Использование:
    uv run python -m scripts.import_mpk path/to/МПК\\ Разработка\\ ПО.xlsx

Скрипт идемпотентен: полностью очищает справочные таблицы МПК
(включая назначенные роли/грейды сотрудников — они обнулятся через ON DELETE SET NULL)
и перезаливает содержимое заново. Существующие оценки (assessments) сносятся
каскадом — на этапе MVP это осознанный выбор.
"""
import asyncio
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select

from app.db import SessionLocal
from app.models.mpk import (
    Assessment,
    AssessmentScore,
    Competency,
    CompetencyCriterion,
    Grade,
    LearningResource,
    ProficiencyLevel,
    Role,
    RoleProfile,
)

GRADES = [
    ("Intern", 0),
    ("Junior", 1),
    ("Junior+", 2),
    ("Middle", 3),
    ("Middle+", 4),
    ("Senior", 5),
    ("Senior+", 6),
]


def split_name_and_desc(raw: str) -> tuple[str, str | None]:
    text = str(raw).replace("\r\n", "\n").strip()
    first, _, rest = text.partition("\n")
    name = first.strip()
    description = rest.strip() if rest.strip() else None
    return name, description


def normalize_header(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip())


def parse_level_header(raw: str) -> tuple[int, str]:
    line = str(raw).strip()
    first, _, rest = line.partition("\n")
    return int(first.strip()), rest.strip()


async def seed_levels(session, ws) -> int:
    count = 0
    for r in range(3, ws.max_row + 1):
        cell = ws.cell(r, 2).value
        if not cell:
            continue
        code, name = parse_level_header(str(cell))
        theory = ws.cell(r, 3).value
        practice = ws.cell(r, 4).value
        comment = ws.cell(r, 5).value
        session.add(
            ProficiencyLevel(
                code=code,
                name=name,
                theory=str(theory).strip() if theory else None,
                practice=str(practice).strip() if practice else None,
                comment=str(comment).strip() if comment and str(comment).strip() != "-" else None,
            )
        )
        count += 1
    return count


async def seed_competencies(session, ws) -> tuple[int, int]:
    comp_count = 0
    crit_count = 0
    current: Competency | None = None
    for r in range(3, ws.max_row + 1):
        code_cell = ws.cell(r, 1).value
        name_cell = ws.cell(r, 2).value
        crit_num = ws.cell(r, 3).value
        crit_desc = ws.cell(r, 4).value

        if code_cell is not None and name_cell:
            name, description = split_name_and_desc(name_cell)
            current = Competency(
                code=int(code_cell),
                name=name,
                description=description,
                sort_order=int(code_cell),
            )
            session.add(current)
            await session.flush()
            comp_count += 1

        if current and crit_num is not None and crit_desc:
            session.add(
                CompetencyCriterion(
                    competency_id=current.id,
                    order_num=int(crit_num),
                    description=str(crit_desc).strip(),
                )
            )
            crit_count += 1
    return comp_count, crit_count


async def seed_grades(session) -> int:
    for code, order in GRADES:
        session.add(Grade(code=code, sort_order=order))
    return len(GRADES)


async def seed_roles_and_profiles(session, ws) -> tuple[int, int]:
    # шапка r2: названия компетенций начиная с col 6
    col_to_comp_name: dict[int, str] = {}
    for c in range(6, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v:
            col_to_comp_name[c] = normalize_header(v)

    # подгружаем компетенции из БД
    comp_rows = (await session.execute(select(Competency))).scalars().all()
    comp_by_name = {c.name: c for c in comp_rows}
    grade_rows = (await session.execute(select(Grade))).scalars().all()
    grade_by_code = {g.code: g for g in grade_rows}

    def resolve_comp(header_name: str) -> Competency | None:
        """Найти компетенцию по заголовку из листа "Профилирование".

        В xlsx заголовок может быть короче полного имени из "Модель ПК".
        Например: header="Разработка Front", в БД="Разработка Front (Angular, React, QT)".
        Делаем prefix-match (case-insensitive) с уникальностью результата.
        """
        if header_name in comp_by_name:
            return comp_by_name[header_name]
        h = header_name.strip().lower()
        candidates = [c for c in comp_rows if c.name.lower().startswith(h)]
        if len(candidates) == 1:
            return candidates[0]
        # ещё одна попытка: подстрока
        candidates = [c for c in comp_rows if h in c.name.lower()]
        if len(candidates) == 1:
            return candidates[0]
        return None

    # первый проход — роли
    roles_map: dict[str, Role] = {}
    for r in range(5, ws.max_row + 1):
        role_name = ws.cell(r, 3).value
        if not role_name:
            continue
        role_name = str(role_name).strip()
        if role_name not in roles_map:
            direction = ws.cell(r, 1).value
            spec = ws.cell(r, 2).value
            role = Role(
                name=role_name,
                direction=str(direction).strip() if direction else None,
                specialization=str(spec).strip() if spec else None,
            )
            session.add(role)
            roles_map[role_name] = role
    await session.flush()

    # второй проход — профили
    profile_count = 0
    for r in range(5, ws.max_row + 1):
        role_name = ws.cell(r, 3).value
        grade_cell = ws.cell(r, 5).value
        if not role_name or not grade_cell:
            continue
        role = roles_map.get(str(role_name).strip())
        grade = grade_by_code.get(str(grade_cell).strip())
        if role is None or grade is None:
            continue
        for col, comp_name in col_to_comp_name.items():
            comp = resolve_comp(comp_name)
            if comp is None:
                continue
            val = ws.cell(r, col).value
            if val is None:
                continue
            try:
                lvl = int(val)
            except (TypeError, ValueError):
                continue
            if not 0 <= lvl <= 5:
                continue
            session.add(
                RoleProfile(
                    role_id=role.id,
                    grade_id=grade.id,
                    competency_id=comp.id,
                    required_level=lvl,
                )
            )
            profile_count += 1
    return len(roles_map), profile_count


async def seed_learning_resources(session, ws) -> int:
    comp_rows = (await session.execute(select(Competency))).scalars().all()
    comp_by_name = {c.name: c for c in comp_rows}

    last_name: str | None = None
    count = 0
    for r in range(4, ws.max_row + 1):
        name_cell = ws.cell(r, 2).value
        if name_cell:
            last_name, _ = split_name_and_desc(name_cell)
        if not last_name:
            continue
        comp = comp_by_name.get(last_name)
        if comp is None:
            continue

        levels: list[int] = []
        for idx, col in enumerate((3, 5, 7, 9), start=1):
            v = ws.cell(r, col).value
            if v and str(v).strip().lower() == "v":
                levels.append(idx)
        if not levels:
            continue

        fmt = ws.cell(r, 11).value
        title = ws.cell(r, 12).value
        provider = ws.cell(r, 13).value
        url = ws.cell(r, 14).value
        evaluation = ws.cell(r, 15).value

        if not title:
            continue

        session.add(
            LearningResource(
                competency_id=comp.id,
                levels=levels,
                format=str(fmt).strip() if fmt else None,
                name=str(title).strip(),
                provider=str(provider).strip() if provider else None,
                url=str(url).strip() if url else None,
                evaluation=str(evaluation).strip() if evaluation else None,
            )
        )
        count += 1
    return count


async def main(path: Path) -> None:
    wb = load_workbook(path, data_only=True)

    required_sheets = [
        "Модель ПК ",
        "Шкала оценки",
        "Новое! Профилирование ПК ",
        "Карта обучения",
    ]
    missing = [s for s in required_sheets if s not in wb.sheetnames]
    if missing:
        print(f"В xlsx не хватает листов: {missing}", file=sys.stderr)
        sys.exit(1)

    async with SessionLocal() as session:
        # порядок важен из-за FK
        await session.execute(delete(AssessmentScore))
        await session.execute(delete(Assessment))
        await session.execute(delete(RoleProfile))
        await session.execute(delete(LearningResource))
        await session.execute(delete(CompetencyCriterion))
        await session.execute(delete(Competency))
        await session.execute(delete(Role))
        await session.execute(delete(Grade))
        await session.execute(delete(ProficiencyLevel))
        await session.flush()

        levels_n = await seed_levels(session, wb["Шкала оценки"])
        comp_n, crit_n = await seed_competencies(session, wb["Модель ПК "])
        grades_n = await seed_grades(session)
        await session.flush()
        roles_n, profiles_n = await seed_roles_and_profiles(
            session, wb["Новое! Профилирование ПК "]
        )
        resources_n = await seed_learning_resources(session, wb["Карта обучения"])

        await session.commit()

    print("Импорт МПК завершён:")
    print(f"  уровни владения:          {levels_n}")
    print(f"  компетенции:              {comp_n}")
    print(f"  индикаторы:               {crit_n}")
    print(f"  грейды:                   {grades_n}")
    print(f"  роли:                     {roles_n}")
    print(f"  профили роль×грейд×комп.: {profiles_n}")
    print(f"  ресурсы обучения:         {resources_n}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("usage: python -m scripts.import_mpk <path_to_xlsx>", file=sys.stderr)
        sys.exit(1)
    p = Path(sys.argv[1]).expanduser()
    if not p.exists():
        print(f"файл не найден: {p}", file=sys.stderr)
        sys.exit(1)
    asyncio.run(main(p))
